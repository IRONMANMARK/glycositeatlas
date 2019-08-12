from openpyxl import load_workbook
import sys
import re

def buildJSONelement(pk,database,accession_uniprotkb,gene_names,protein_names,glycosylation_location,
                      glycosites_containing_peptides,glycosite_aas,identification_resources,years,references):
    je = "  {\n"
    je += "    \"model\": \"glycosites.GlycositeRecord\",\n"
    je += "    \"pk\": {0:d},\n".format(pk)
    je += "    \"fields\": {\n"
    je += "      \"database\": \"{0:s}\",\n".format(database)
    je += "      \"accession_uniprotkb\": \"{0:s}\",\n".format(accession_uniprotkb)
    je += "      \"gene_names\": \"{0:s}\",\n".format(gene_names)
    je += "      \"protein_names\": \"{0:s}\",\n".format(protein_names)
    je += "      \"glycosylation_location\": \"{0:d}\",\n".format(glycosylation_location)
    je += "      \"glycosites_containing_peptides\": \"{0:s}\",\n".format(glycosites_containing_peptides)
    je += "      \"glycosite_aas\": \"{0:s}\",\n".format(glycosite_aas)
    je += "      \"identification_resources\": \"{0:s}\",\n".format(str(identification_resources.encode("GBK",'ignore')))
    je += "      \"years\": \"{0:s}\",\n".format(years)
    je += "      \"references\": \"{0:s}\"\n".format(references)
    je += "    }\n"
    je += "  }"
    return je

def main():
    # input_xlsx = '20151118_GlycositeDB_Human.xlsx'
    input_xlsx = 'HumanAll.xlsx'
    output_jason = re.sub(".xlsx$",".json",input_xlsx)

    # wb = load_workbook('20151118_GlycositeDB_Human.xlsx')
    wb = load_workbook(input_xlsx)
    # print(wb.get_sheet_names())
    # ws = wb['GlycositeDB']
    ws = wb['Glycosite and glycoprotein DB']

    of = open(output_jason,"w")
    of.write("[\n")
    pk = 0
    rows_count = len(ws.rows)
    for row in ws.rows:
        if pk == 0:
            pk += 1
            continue
        database = row[0].value
        accession_uniprotkb = row[1].value
        gene_names = row[2].value
        protein_names = row[3].value
        glycosylation_location = int(row[4].value)
        glycosites_containing_peptides = row[5].value
        glycosite_aas = row[6].value
        identification_resources = str(row[7].value)
        years = str(row[8].value)
        references = str(row[9].value)
        je = buildJSONelement(pk,database,accession_uniprotkb,gene_names,protein_names,glycosylation_location,
                          glycosites_containing_peptides,glycosite_aas,identification_resources,years,references)
        if pk == rows_count-1:
            of.write(je+"\n")
        else:
            of.write(je)
            of.write(",")
            of.write("\n")
        pk += 1
    of.write("]\n")
    of.close()

if __name__ == "__main__":
    main()